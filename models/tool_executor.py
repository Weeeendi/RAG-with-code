from typing import Dict, List, Any, Callable, Optional
from dataclasses import dataclass
import os


@dataclass
class ToolResult:
    success: bool
    data: Any
    error: Optional[str] = None


class ToolRegistry:
    _tools: Dict[str, Callable] = {}
    _extensions: Dict[str, List[tuple]] = {}

    @classmethod
    def register(cls, name: str, handler: Callable, extensions: List[str] = None, priority: int = 0):
        cls._tools[name] = handler
        if extensions:
            for ext in extensions:
                ext_lower = ext.lower()
                if ext_lower not in cls._extensions:
                    cls._extensions[ext_lower] = []
                inserted = False
                for i, (p, _) in enumerate(cls._extensions[ext_lower]):
                    if priority > p:
                        cls._extensions[ext_lower].insert(i, (priority, name))
                        inserted = True
                        break
                if not inserted:
                    cls._extensions[ext_lower].append((priority, name))

    @classmethod
    def get(cls, name: str) -> Optional[Callable]:
        return cls._tools.get(name)

    @classmethod
    def get_by_extension(cls, ext: str) -> Optional[str]:
        ext_list = cls._extensions.get(ext.lower(), [])
        return ext_list[0][1] if ext_list else None

    @classmethod
    def list_by_extension(cls, ext: str) -> List[str]:
        return [name for _, name in cls._extensions.get(ext.lower(), [])]

    @classmethod
    def list_tools(cls) -> List[str]:
        return list(cls._tools.keys())


class ToolExecutor:
    def __init__(self, image_store_dir: str = "knowledge_base/raw/images",
                 parsed_dir: str = "knowledge_base/parsed"):
        self.registry = ToolRegistry
        self.image_store_dir = image_store_dir
        self.parsed_dir = parsed_dir
        os.makedirs(image_store_dir, exist_ok=True)
        os.makedirs(parsed_dir, exist_ok=True)
        self._register_default_tools()

    def _get_parsed_cache_path(self, file_path: str, tool_name: str = None) -> str:
        """Get the cache file path for a parsed source file."""
        import hashlib
        rel_path = os.path.relpath(file_path, os.getcwd())
        # 缓存 key 包含工具名称，这样不同工具不会共享缓存
        cache_key_base = f"{rel_path}_{tool_name or 'default'}"
        cache_key = hashlib.md5(cache_key_base.encode()).hexdigest()[:12]
        cache_dir = os.path.join(self.parsed_dir, cache_key)
        return os.path.join(cache_dir, f"{os.path.basename(file_path)}_{tool_name or 'default'}.cache.json")

    def _get_file_hash(self, file_path: str) -> str:
        """Compute hash of source file for change detection."""
        import hashlib
        with open(file_path, 'rb') as f:
            return hashlib.md5(f.read()).hexdigest()

    def _load_cached_parse(self, file_path: str, tool_name: str = None) -> Optional[dict]:
        """Load cached parse result if file hasn't changed."""
        cache_path = self._get_parsed_cache_path(file_path, tool_name)
        if not os.path.exists(cache_path):
            return None

        try:
            import json
            with open(cache_path, 'r', encoding='utf-8') as f:
                cache = json.load(f)

            current_hash = self._get_file_hash(file_path)
            if cache.get('file_hash') != current_hash:
                return None

            return cache
        except Exception:
            return None

    def _save_cached_parse(self, file_path: str, data: Any, tool_name: str = None) -> None:
        """Save parse result to cache."""
        import json
        import hashlib

        cache_path = self._get_parsed_cache_path(file_path, tool_name)
        cache_dir = os.path.dirname(cache_path)
        os.makedirs(cache_dir, exist_ok=True)

        cache = {
            'file_hash': self._get_file_hash(file_path),
            'data': data
        }

        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)

    def _register_default_tools(self):
        # priority 数值越大越优先，pdfplumber(10) > paddleocr(0)
        # 主线路: pdfplumber → 失败则 paddleocr OCR 回退
        self.registry.register('pdf', self._parse_pdf, ['.pdf'], priority=10)
        self.registry.register('paddleocr', self._parse_pdf_paddleocr, ['.pdf'], priority=0)
        self.registry.register('docx', self._parse_docx, ['.docx', '.doc'], priority=0)
        self.registry.register('excel', self._parse_excel, ['.xlsx', '.xls'], priority=0)
        self.registry.register('txt', self._parse_txt, ['.txt'], priority=0)
        self.registry.register('md', self._parse_md, ['.md'], priority=0)

    def _parse_pdf(self, file_path: str, **kwargs) -> ToolResult:
        try:
            import pdfplumber
            import hashlib
            pages = []
            pdf_hash = hashlib.md5(file_path.encode()).hexdigest()[:8]

            with pdfplumber.open(file_path) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    page_data = {'page': page_num, 'text': '', 'tables': [], 'images': [], 'source': file_path}

                    text = page.extract_text()
                    if text:
                        page_data['text'] = text

                    tables = page.extract_tables()
                    if tables:
                        for table in tables:
                            if table and len(table) > 0:
                                cleaned_table = []
                                for row in table:
                                    cleaned_row = [str(cell).strip() if cell else '' for cell in row]
                                    if any(cell for cell in cleaned_row):
                                        cleaned_table.append(cleaned_row)
                                if cleaned_table:
                                    page_data['tables'].append(cleaned_table)

                    images = page.images
                    if images:
                        saved_image_paths = self._extract_images_by_pymupdf(file_path, page_num, pdf_hash)
                        if saved_image_paths:
                            page_data['images'] = saved_image_paths

                    if page_data['text'] or page_data['tables'] or page_data['images']:
                        pages.append(page_data)
            return ToolResult(success=True, data=pages)
        except Exception as e:
            return ToolResult(success=False, data=None, error=str(e))

    def _extract_images_by_pymupdf(self, pdf_path: str, page_num: int, pdf_hash: str) -> List[Dict]:
        try:
            import fitz
            saved = []
            doc = fitz.open(pdf_path)
            if page_num > len(doc):
                return saved
            page = doc[page_num - 1]
            image_list = page.get_images(full=True)
            for img_idx, img in enumerate(image_list):
                xref = img[0]
                base_image = doc.extract_image(xref)
                img_data = base_image.get('image')
                if img_data:
                    img_ext = base_image.get('ext', 'jpeg')
                    img_name = f"{pdf_hash}_p{page_num}_i{img_idx}.{img_ext}"
                    img_path = os.path.join(self.image_store_dir, img_name)
                    try:
                        with open(img_path, 'wb') as f:
                            f.write(img_data)
                        saved.append({
                            'path': img_path,
                            'page': page_num,
                            'index': img_idx,
                            'width': base_image.get('width'),
                            'height': base_image.get('height')
                        })
                    except Exception:
                        pass
            doc.close()
            return saved
        except Exception:
            return []

    def _guess_image_ext(self, format_str: str) -> str:
        format_lower = str(format_str).lower()
        if 'png' in format_lower:
            return '.png'
        elif 'gif' in format_lower:
            return '.gif'
        elif 'bmp' in format_lower:
            return '.bmp'
        return '.jpg'

    def _parse_docx(self, file_path: str, **kwargs) -> ToolResult:
        try:
            from docx import Document
            doc = Document(file_path)
            paragraphs = [para.text for para in doc.paragraphs]
            return ToolResult(success=True, data=paragraphs)
        except Exception as e:
            return ToolResult(success=False, data=None, error=str(e))

    def _parse_excel(self, file_path: str, **kwargs) -> ToolResult:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            sheets = {}
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        rows.append([cell for cell in row])
                sheets[sheet_name] = rows
            return ToolResult(success=True, data=sheets)
        except Exception as e:
            return ToolResult(success=False, data=None, error=str(e))

    def _parse_txt(self, file_path: str, **kwargs) -> ToolResult:
        try:
            from models.utils.text_cleaner import read_file_with_encoding, looks_corrupted
            encoding, content = read_file_with_encoding(file_path)
            if content.strip():
                if looks_corrupted(content, threshold=0.05):
                    import logging
                    logging.warning(f"⚠️ possible corrupted text in {file_path} — first 100 chars: {repr(content[:100])}")
                return ToolResult(success=True, data=content)
            return ToolResult(success=False, data=None, error="文件内容为空")
        except Exception as e:
            return ToolResult(success=False, data=None, error=str(e))

    def _parse_md(self, file_path: str, **kwargs) -> ToolResult:
        try:
            import re
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            sections = []
            tables = []
            current_section = {'page': 1, 'text': '', 'tables': []}
            lines = content.split('\n')
            section_counter = 1

            table_buffer = []
            in_table = False

            for line in lines:
                table_match = re.match(r'^\|(.+)\|$', line)
                if table_match:
                    row_cells = [c.strip() for c in table_match.group(1).split('|')]
                    if not any(re.match(r'^[-:]+$', cell) for cell in row_cells):
                        if not in_table:
                            in_table = True
                            table_buffer = []
                        table_buffer.append(row_cells)
                    elif in_table and table_buffer:
                        tables.append(table_buffer)
                        table_buffer = []
                        in_table = False
                else:
                    if in_table and table_buffer:
                        tables.append(table_buffer)
                        table_buffer = []
                        in_table = False

                    if re.match(r'^#{1,6}\s+', line):
                        if current_section['text'].strip():
                            sections.append(current_section)
                        current_section = {'page': section_counter, 'text': line + '\n', 'tables': []}
                        section_counter += 1
                    else:
                        current_section['text'] += line + '\n'

            if in_table and table_buffer:
                tables.append(table_buffer)

            if current_section['text'].strip():
                sections.append(current_section)

            for section in sections:
                section['tables'] = []
                for table in tables:
                    table_text = self._format_md_table(table)
                    section['tables'].append(table_text)
                    section['text'] += '\n' + table_text

            if not sections:
                sections.append({'page': 1, 'text': content, 'tables': []})
            return ToolResult(success=True, data=sections)
        except Exception as e:
            return ToolResult(success=False, data=None, error=str(e))

    def _parse_pdf_paddleocr(self, file_path: str, **kwargs) -> ToolResult:
        try:
            # Check cache first (use tool_name='paddleocr' for separate cache)
            cached = self._load_cached_parse(file_path, tool_name='paddleocr')
            if cached is not None:
                print(f"[PaddleOCR] Cache hit: {os.path.basename(file_path)}")
                return ToolResult(success=True, data=cached['data'])

            from models.paddleocr_parser import EnhancedPDFParser
            use_llm = kwargs.get('use_llm_reorganize', True)
            output_dir = kwargs.get('output_dir')
            parser = EnhancedPDFParser(output_dir or os.path.join(os.path.dirname(file_path), "paddleocr_output"))
            blocks = parser.parse_and_process(file_path, use_llm_reorganize=use_llm)

            # Save to cache with tool_name
            self._save_cached_parse(file_path, blocks, tool_name='paddleocr')

            return ToolResult(success=True, data=blocks)
        except Exception as e:
            print(f"[PaddleOCR] Failed, falling back to pdfplumber: {e}")
            return self._parse_pdf(file_path, **kwargs)

    def _format_md_table(self, table: list) -> str:
        if not table:
            return ''
        lines = []
        for row in table:
            lines.append('| ' + ' | '.join(str(cell) for cell in row) + ' |')
        return '\n'.join(lines)

    def execute(self, file_path: str, tool_name: str = None, **kwargs) -> ToolResult:
        if tool_name is None:
            ext = os.path.splitext(file_path)[1]
            tool_names = self.registry.list_by_extension(ext)
            if not tool_names:
                return ToolResult(success=False, data=None, error=f"No tool found for extension: {ext}")

            # 按优先级尝试每个工具
            errors = []
            for tn in tool_names:
                tool = self.registry.get(tn)
                if tool:
                    result = tool(file_path, **kwargs)
                    if result.success:
                        return result
                    errors.append(f"{tn}: {result.error}")
            return ToolResult(success=False, data=None, error="; ".join(errors))
        else:
            tool = self.registry.get(tool_name)
            if tool is None:
                return ToolResult(success=False, data=None, error=f"Tool not found: {tool_name}")
            return tool(file_path, **kwargs)

    def execute_by_extension(self, file_path: str, **kwargs) -> ToolResult:
        return self.execute(file_path, **kwargs)

    def register_tool(self, name: str, handler: Callable, extensions: List[str] = None):
        self.registry.register(name, handler, extensions)
